"""_extractHMI.py
Extract HMI-accessible variables from TIA Portal DB/IDB XML and .db exports.
Outputs Excel with columns matching TIA Portal DB monitor view.
Resolves UDT structures recursively from PLC data type definitions.

Dependencies:
    pip install pyyaml openpyxl

If you get:
    ModuleNotFoundError: No module named 'yaml'
install the missing package with:
    pip install pyyaml

Usage:
    python Tools\\_extractHMI.py                           # auto-detect TiaExport in cwd
    python Tools\\_extractHMI.py <plc_base_dir>             # explicit PLC folder path
    python Tools\\_extractHMI.py <plc_base_dir> -o out.xlsx  # custom output file
    python Tools\\_extractHMI.py --help
"""

import argparse
import os
import re
import sys
import yaml
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === Configuration ===

NS = "http://www.siemens.com/automation/Openness/SW/Interface/v5"

ARRAY_EXPAND_LIMIT = 64

BLOCK_TYPES_XML = {
    "SW.Blocks.GlobalDB": "GlobalDB",
    "SW.Blocks.InstanceDB": "InstanceDB",
}

DEFAULT_START_VALUES = {
    "Bool": "FALSE",
    "Byte": "16#0", "Word": "16#0", "DWord": "16#0", "LWord": "16#0",
    "SInt": "0", "USInt": "0", "Int": "0", "UInt": "0",
    "DInt": "0", "UDInt": "0", "LInt": "0", "ULInt": "0",
    "Real": "0.0", "LReal": "0.0",
    "Char": "' '", "WChar": "' '",
    "String": "''", "WString": "''",
    "Time": "T#0ms", "LTime": "LT#0ns",
    "Date": "D#1990-01-01",
    "Time_Of_Day": "TOD#00:00:00",
    "Date_And_Time": "DT#1990-01-01-00:00:00",
    "DTL": "DTL#1970-01-01-00:00:00.0",
}

EXCEL_HEADERS = [
    "Name",
    "Data type",
    "Start value",
    "Retain",
    "Accessible from HMI/OPC UA",
    "Writable from HMI/OPC",
    "Visible in HMI engineering",
    "Setpoint",
    "Comment",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  Data Structures
# ═══════════════════════════════════════════════════════════════════════════════

class MemberInfo:
    """Cached member info from UDT or FB interface definition."""
    __slots__ = ("name", "datatype", "attrs", "comment", "children")

    def __init__(self, name, datatype, attrs, comment, children=None):
        self.name = name
        self.datatype = datatype
        self.attrs = attrs          # {"ExternalAccessible": bool, ...}
        self.comment = comment      # str
        self.children = children or {}  # {child_name: MemberInfo}


# ═══════════════════════════════════════════════════════════════════════════════
#  Utility Functions
# ═══════════════════════════════════════════════════════════════════════════════

def find_xml_files(base_dir):
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith(".xml"):
                yield os.path.join(root, f)


def find_db_files(base_dir):
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith(".db"):
                yield os.path.join(root, f)


def normalize_udt_name(datatype):
    """Strip quotes from UDT reference:  \"Name\" → Name,  &quot;Name&quot; → Name."""
    dt = datatype.strip()
    if dt.startswith('"') and dt.endswith('"') and len(dt) > 2:
        return dt[1:-1]
    if dt.startswith('&quot;') and dt.endswith('&quot;') and len(dt) > 12:
        return dt[6:-6]
    return None


def is_udt_reference(datatype):
    return normalize_udt_name(datatype) is not None


def parse_array_type(datatype):
    """Parse Array[lo..hi] of ElementType → (lo, hi, element_type) or None."""
    m = re.match(r'Array\[(\d+)\.\.(\d+)\]\s+of\s+(.+)', datatype, re.IGNORECASE)
    if m:
        return int(m.group(1)), int(m.group(2)), m.group(3).strip()
    return None


def format_datatype(datatype):
    """Normalize UDT quotes for display: &quot;X&quot; → \"X\"."""
    udt = normalize_udt_name(datatype)
    if udt:
        return f'"{udt}"'
    return datatype


def get_default_start_value(datatype):
    dt = datatype.strip()
    udt = normalize_udt_name(dt)
    if udt:
        return ""
    if dt in DEFAULT_START_VALUES:
        return DEFAULT_START_VALUES[dt]
    return ""


def _local_tag(elem):
    """Get local tag name without namespace."""
    tag = elem.tag
    return tag.split("}")[-1] if "}" in tag else tag


# ═══════════════════════════════════════════════════════════════════════════════
#  XML Parsing Helpers (namespace-agnostic)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_boolean_attrs(member_elem):
    """Extract ExternalAccessible/Visible/Writable/SetPoint from a Member."""
    attrs = {}
    for child in member_elem:
        if _local_tag(child) == "AttributeList":
            for ba in child:
                if _local_tag(ba) == "BooleanAttribute":
                    name = ba.get("Name")
                    if name in ("ExternalAccessible", "ExternalVisible",
                                "ExternalWritable", "SetPoint"):
                        attrs[name] = (ba.text or "").strip().lower() == "true"
            break
    return attrs


def get_comment_text(elem):
    """Extract en-US comment text from a Member element."""
    for child in elem:
        if _local_tag(child) == "Comment":
            best = ""
            for mlt in child:
                text = (mlt.text or "").strip()
                if mlt.get("Lang") == "en-US" and text:
                    return text
                if text and not best:
                    best = text
            return best
    return ""


def get_start_value_text(elem):
    """Extract StartValue text from a Member element."""
    for child in elem:
        if _local_tag(child) == "StartValue":
            return (child.text or "").strip()
    return ""


def get_child_member_elems(elem):
    """Get child Member elements: direct Members + Sections/Section/Members."""
    children = []
    for child in elem:
        lt = _local_tag(child)
        if lt == "Member":
            children.append(child)
        elif lt == "Sections":
            for sec in child:
                if _local_tag(sec) == "Section":
                    for m in sec:
                        if _local_tag(m) == "Member":
                            children.append(m)
    return children


# ═══════════════════════════════════════════════════════════════════════════════
#  UDT Cache
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_udt_member(member_elem):
    """Parse a Member element from a UDT into MemberInfo (recursive)."""
    name = member_elem.get("Name", "")
    datatype = member_elem.get("Datatype", "")
    attrs = extract_boolean_attrs(member_elem)
    comment = get_comment_text(member_elem)
    children = {}

    # Collect child members (direct for Struct, Sections for UDT inline expansion)
    for child_elem in get_child_member_elems(member_elem):
        info = _parse_udt_member(child_elem)
        children[info.name] = info

    return MemberInfo(name, datatype, attrs, comment, children)


def build_udt_cache(udt_dir):
    """Build {udt_name: {member_name: MemberInfo}} from PLC data type XMLs."""
    cache = {}
    for filepath in find_xml_files(udt_dir):
        try:
            tree = ET.parse(filepath)
        except ET.ParseError:
            continue
        root = tree.getroot()
        struct_elem = root.find(".//SW.Types.PlcStruct")
        if struct_elem is None:
            continue
        attr = struct_elem.find("AttributeList")
        if attr is None:
            continue
        name_elem = attr.find("Name")
        if name_elem is None or not name_elem.text:
            continue
        udt_name = name_elem.text.strip()

        iface_sections = attr.find(f"Interface/{{{NS}}}Sections")
        if iface_sections is None:
            continue
        for section in iface_sections.findall(f"{{{NS}}}Section"):
            if section.get("Name") == "None":
                members = {}
                for m in section.findall(f"{{{NS}}}Member"):
                    info = _parse_udt_member(m)
                    members[info.name] = info
                cache[udt_name] = members
                break
    return cache


# ═══════════════════════════════════════════════════════════════════════════════
#  FB Interface Cache (for top-level comments in IDBs)
# ═══════════════════════════════════════════════════════════════════════════════

# --- .s7dcl + .s7res Parser ---

_RE_BLOCK_DECL = re.compile(
    r'(FUNCTION_BLOCK|FUNCTION|ORGANIZATION_BLOCK)\s+"([^"]+)"'
)
_RE_VAR_SECTION = re.compile(
    r'^\s*(VAR_INPUT|VAR_OUTPUT|VAR_IN_OUT|VAR_TEMP|VAR\s+RETAIN|VAR)\b',
    re.IGNORECASE,
)
_RE_END_VAR = re.compile(r'^\s*END_VAR\b', re.IGNORECASE)
_RE_S7_ATTR = re.compile(r'S7_MLC\s*:=\s*"([^"]+)"')
_RE_MEMBER_S7DCL = re.compile(r'^\s+(\w+)\s*:')

_SECTION_MAP = {
    "VAR_INPUT": "Input",
    "VAR_OUTPUT": "Output",
    "VAR_IN_OUT": "InOut",
    "VAR_TEMP": "Temp",
    "VAR RETAIN": "Static",
    "VAR": "Static",
}


def _parse_s7res(res_path):
    """Parse .s7res YAML file → {mlc_id: en-US text}."""
    result = {}
    try:
        with open(res_path, "r", encoding="utf-8-sig") as f:
            data = yaml.safe_load(f)
    except Exception:
        return result
    if not data or "MultiLingualTexts" not in data:
        return result
    for entry in data["MultiLingualTexts"]:
        mid = entry.get("id", "")
        text = entry.get("en-US", "") or entry.get("pl-PL", "")
        if mid and text:
            result[mid] = str(text).strip()
    return result


def _parse_s7dcl_members(dcl_path, mlc_texts):
    """Parse .s7dcl file → {section_name: {member_name: MemberInfo}}."""
    try:
        with open(dcl_path, "r", encoding="utf-8-sig") as f:
            lines = f.readlines()
    except Exception:
        return None, {}

    block_name = None
    fb_sections = {}
    current_section = None
    pending_mlc = None

    for line in lines:
        # Block name
        m = _RE_BLOCK_DECL.search(line)
        if m:
            block_name = m.group(2)
            continue

        # VAR section start
        m = _RE_VAR_SECTION.match(line)
        if m:
            raw = m.group(1).strip()
            current_section = _SECTION_MAP.get(raw, "Static")
            pending_mlc = None
            continue

        # END_VAR
        if _RE_END_VAR.match(line):
            current_section = None
            pending_mlc = None
            continue

        if current_section is None:
            continue

        # Attribute block { ... S7_MLC := "MLC_xxx" ... }
        mlc_m = _RE_S7_ATTR.search(line)
        if mlc_m:
            pending_mlc = mlc_m.group(1)
            # Don't continue — the member might be on the same line or next

        # Member declaration
        mem_m = _RE_MEMBER_S7DCL.match(line)
        if mem_m:
            member_name = mem_m.group(1)
            comment = mlc_texts.get(pending_mlc, "") if pending_mlc else ""
            info = MemberInfo(member_name, "", {}, comment)
            fb_sections.setdefault(current_section, {})[member_name] = info
            pending_mlc = None

    return block_name, fb_sections


def build_fb_cache(prog_dir):
    """Build {fb_name: {section_name: {member_name: MemberInfo}}}.
    Sources: FB/FC/OB XML files + .s7dcl/.s7res file pairs."""
    cache = {}

    # 1) Parse XML FB/FC/OB files
    for filepath in find_xml_files(prog_dir):
        try:
            tree = ET.parse(filepath)
        except ET.ParseError:
            continue
        root = tree.getroot()

        block_elem = None
        for tag in ("SW.Blocks.FB", "SW.Blocks.FC", "SW.Blocks.OB"):
            block_elem = root.find(f".//{tag}")
            if block_elem is not None:
                break
        if block_elem is None:
            continue

        attr = block_elem.find("AttributeList")
        if attr is None:
            continue
        name_elem = attr.find("Name")
        if name_elem is None or not name_elem.text:
            continue
        fb_name = name_elem.text.strip()

        iface_sections = attr.find(f"Interface/{{{NS}}}Sections")
        if iface_sections is None:
            continue

        fb_sections = {}
        for section in iface_sections.findall(f"{{{NS}}}Section"):
            section_name = section.get("Name", "")
            members = {}
            for m in section.findall(f"{{{NS}}}Member"):
                info = _parse_udt_member(m)
                members[info.name] = info
            if members:
                fb_sections[section_name] = members
        if fb_sections:
            cache[fb_name] = fb_sections

    # 2) Parse .s7dcl + .s7res file pairs
    xml_fb_count = len(cache)
    for root_dir, _, files in os.walk(prog_dir):
        for f in files:
            if not f.lower().endswith(".s7dcl"):
                continue
            dcl_path = os.path.join(root_dir, f)
            res_path = os.path.splitext(dcl_path)[0] + ".s7res"

            mlc_texts = _parse_s7res(res_path) if os.path.isfile(res_path) else {}
            block_name, fb_sections = _parse_s7dcl_members(dcl_path, mlc_texts)

            if block_name and fb_sections and block_name not in cache:
                cache[block_name] = fb_sections

    return cache, xml_fb_count


# ═══════════════════════════════════════════════════════════════════════════════
#  Attribute Resolution
# ═══════════════════════════════════════════════════════════════════════════════

ATTR_KEYS = ("ExternalAccessible", "ExternalVisible", "ExternalWritable", "SetPoint")
ATTR_DEFAULTS = {
    "ExternalAccessible": True,
    "ExternalVisible": True,
    "ExternalWritable": True,
    "SetPoint": False,
}


def resolve_attrs(member_elem, udt_context):
    """Resolve boolean attributes: own XML → UDT context → defaults."""
    own = extract_boolean_attrs(member_elem)
    result = {}
    for key in ATTR_KEYS:
        if key in own:
            result[key] = own[key]
        elif udt_context and key in udt_context.attrs:
            result[key] = udt_context.attrs[key]
        else:
            result[key] = ATTR_DEFAULTS[key]
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  Main Walk: IDB/GlobalDB XML Members
# ═══════════════════════════════════════════════════════════════════════════════

def walk_member(elem, path_parts, parent_remanence, udt_context, udt_cache, results,
                depth=0):
    """Recursively walk a DB member element and collect HMI-accessible rows."""
    if depth > 30:
        return

    name = elem.get("Name", "")
    datatype = elem.get("Datatype", "")
    current_path = path_parts + [name]

    # --- Resolve attributes ---
    attrs = resolve_attrs(elem, udt_context)

    if not attrs["ExternalAccessible"]:
        return  # skip entire subtree

    # --- Comment ---
    comment = get_comment_text(elem)
    if not comment and udt_context:
        comment = udt_context.comment

    # --- Remanence ---
    rem = elem.get("Remanence", "")
    remanence = rem if rem else parent_remanence

    # --- Start value ---
    start_value = get_start_value_text(elem)

    # --- Display datatype ---
    display_dt = format_datatype(datatype)

    # --- Add row ---
    results.append({
        "Name": ".".join(current_path),
        "Data type": display_dt,
        "Start value": start_value,
        "Retain": remanence == "Retain",
        "Accessible from HMI/OPC UA": attrs["ExternalAccessible"],
        "Writable from HMI/OPC": attrs["ExternalWritable"],
        "Visible in HMI engineering": attrs["ExternalVisible"],
        "Setpoint": attrs["SetPoint"],
        "Comment": comment,
    })

    # --- Determine child UDT context map ---
    child_udt_map = _resolve_child_udt_map(datatype, udt_context, udt_cache)

    # --- Collect child elements from XML ---
    child_elems = get_child_member_elems(elem)

    # --- Array expansion (virtual child rows if no XML children) ---
    arr = parse_array_type(datatype)
    if arr and not child_elems:
        lo, hi, elem_type = arr
        count = hi - lo + 1
        if count <= ARRAY_EXPAND_LIMIT:
            _expand_array(current_path, lo, hi, elem_type, remanence, attrs,
                          udt_cache, results, depth)

    # --- Recurse into children ---
    for child_elem in child_elems:
        child_name = child_elem.get("Name", "")
        child_udt = child_udt_map.get(child_name) if child_udt_map else None
        walk_member(child_elem, current_path, remanence, child_udt,
                    udt_cache, results, depth + 1)


def _resolve_child_udt_map(datatype, udt_context, udt_cache):
    """Determine the UDT member map to use for children."""
    udt_name = normalize_udt_name(datatype)
    if udt_name and udt_name in udt_cache:
        return udt_cache[udt_name]

    if datatype == "Struct" and udt_context and udt_context.children:
        return udt_context.children

    # Current member might not have UDT in datatype, but udt_context knows the type
    if udt_context:
        ctx_udt = normalize_udt_name(udt_context.datatype)
        if ctx_udt and ctx_udt in udt_cache:
            return udt_cache[ctx_udt]

    return None


def _expand_array(parent_path, lo, hi, elem_type, remanence, parent_attrs,
                  udt_cache, results, depth):
    """Create virtual rows for array elements."""
    display_et = format_datatype(elem_type)
    default_sv = get_default_start_value(elem_type)
    udt_name = normalize_udt_name(elem_type)
    elem_udt_map = udt_cache.get(udt_name) if udt_name else None

    parent_name = parent_path[-1]
    base_path = parent_path[:-1]

    for i in range(lo, hi + 1):
        idx_name = f"{parent_name}[{i}]"
        idx_path = base_path + [idx_name]

        results.append({
            "Name": ".".join(idx_path),
            "Data type": display_et,
            "Start value": default_sv,
            "Retain": remanence == "Retain",
            "Accessible from HMI/OPC UA": parent_attrs["ExternalAccessible"],
            "Writable from HMI/OPC": parent_attrs["ExternalWritable"],
            "Visible in HMI engineering": parent_attrs["ExternalVisible"],
            "Setpoint": parent_attrs["SetPoint"],
            "Comment": "",
        })

        # If element type is UDT, expand members
        if elem_udt_map:
            for m_info in elem_udt_map.values():
                _walk_udt_from_cache(m_info, idx_path, remanence, udt_cache,
                                     results, depth + 1)


def _walk_udt_from_cache(m_info, path_parts, remanence, udt_cache, results, depth):
    """Walk a UDT member from cache (for array or pure-UDT expansion without IDB XML)."""
    if depth > 30:
        return

    current_path = path_parts + [m_info.name]
    attrs = m_info.attrs if m_info.attrs else ATTR_DEFAULTS.copy()

    if not attrs.get("ExternalAccessible", True):
        return

    display_dt = format_datatype(m_info.datatype)

    results.append({
        "Name": ".".join(current_path),
        "Data type": display_dt,
        "Start value": "",
        "Retain": remanence == "Retain",
        "Accessible from HMI/OPC UA": attrs.get("ExternalAccessible", True),
        "Writable from HMI/OPC": attrs.get("ExternalWritable", True),
        "Visible in HMI engineering": attrs.get("ExternalVisible", True),
        "Setpoint": attrs.get("SetPoint", False),
        "Comment": m_info.comment,
    })

    # Determine children
    child_map = None
    udt_name = normalize_udt_name(m_info.datatype)
    if udt_name and udt_name in udt_cache:
        child_map = udt_cache[udt_name]
    elif m_info.children:
        child_map = m_info.children

    if child_map:
        for child_info in child_map.values():
            _walk_udt_from_cache(child_info, current_path, remanence,
                                 udt_cache, results, depth + 1)


# ═══════════════════════════════════════════════════════════════════════════════
#  Process XML File (InstanceDB / GlobalDB)
# ═══════════════════════════════════════════════════════════════════════════════

def process_xml_file(filepath, udt_cache, fb_cache):
    """Process DB/IDB XML file; return list of result dicts."""
    try:
        tree = ET.parse(filepath)
    except ET.ParseError:
        return []

    root = tree.getroot()

    for block_tag in BLOCK_TYPES_XML:
        block_elem = root.find(f".//{block_tag}")
        if block_elem is None:
            continue

        attr = block_elem.find("AttributeList")
        if attr is None:
            return []

        name_elem = attr.find("Name")
        block_name = (name_elem.text or "").strip() if name_elem is not None else ""
        if not block_name:
            return []

        instance_of = ""
        if BLOCK_TYPES_XML[block_tag] == "InstanceDB":
            iof = attr.find("InstanceOfName")
            if iof is not None and iof.text:
                instance_of = iof.text.strip()

        iface_sections = attr.find(f"Interface/{{{NS}}}Sections")
        if iface_sections is None:
            return []

        # FB interface for comment fallback
        fb_sections = fb_cache.get(instance_of, {}) if instance_of else {}

        results = []
        for section in iface_sections.findall(f"{{{NS}}}Section"):
            section_name = section.get("Name", "")
            fb_members = fb_sections.get(section_name, {})

            for member in section.findall(f"{{{NS}}}Member"):
                member_name = member.get("Name", "")

                # Build UDT context for this top-level member
                udt_context = fb_members.get(member_name)

                walk_member(member, [block_name], "NonRetain", udt_context,
                            udt_cache, results)

        return results

    return []


# ═══════════════════════════════════════════════════════════════════════════════
#  Process .db Text File (Global Data Block)
# ═══════════════════════════════════════════════════════════════════════════════

_RE_DB_HEADER = re.compile(r'\s*DATA_BLOCK\s+"([^"]+)"')
_RE_MEMBER_LINE = re.compile(
    r'(\s+)'                       # leading whitespace (indent)
    r'(\w+)\s*'                    # member name
    r'(?:\{([^}]*)\})?\s*'         # optional {attrs}
    r':\s*'                        # colon
    r'(.+)$'                       # rest: type [;] [// comment]
)


def _parse_db_attrs(attrs_str):
    """Parse {key:='val'; ...} from .db file into attrs dict."""
    attrs = dict(ATTR_DEFAULTS)
    if not attrs_str:
        return attrs
    for pair in attrs_str.split(";"):
        pair = pair.strip()
        if not pair:
            continue
        m = re.match(r"(\w+)\s*:=\s*'([^']*)'", pair)
        if m:
            key, val = m.group(1), m.group(2).strip().lower() == "true"
            if key == "ExternalAccessible":
                attrs["ExternalAccessible"] = val
            elif key == "ExternalVisible":
                attrs["ExternalVisible"] = val
            elif key == "ExternalWritable":
                attrs["ExternalWritable"] = val
            elif key == "S7_SetPoint":
                attrs["SetPoint"] = val
    return attrs


def process_db_file(filepath, udt_cache):
    """Parse a .db text file and extract HMI rows."""
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            content = f.read()
    except (IOError, UnicodeDecodeError):
        return []

    m = _RE_DB_HEADER.match(content)
    if not m:
        return []
    block_name = m.group(1)

    if re.search(r'^\s*NON_RETAIN\b', content, re.MULTILINE):
        base_remanence = "NonRetain"
    elif re.search(r'^\s*RETAIN\b', content, re.MULTILINE):
        base_remanence = "Retain"
    else:
        base_remanence = "NonRetain"

    lines = content.split("\n")
    results = []
    path_stack = []   # [(name, attrs)]
    in_var = False
    in_begin = False
    skip_depth = 0    # nesting levels of inaccessible Structs to skip

    for line in lines:
        stripped = line.strip()

        if stripped.startswith("END_DATA_BLOCK"):
            break
        if stripped.startswith("BEGIN"):
            in_var = False
            in_begin = True
            continue
        if in_begin:
            continue

        if stripped in ("VAR", "STRUCT") and not in_var:
            in_var = True
            continue
        if stripped.startswith("END_VAR") or stripped == "END_STRUCT;":
            if skip_depth > 0:
                skip_depth -= 1
            elif path_stack:
                path_stack.pop()
            else:
                in_var = False
            continue

        if not in_var:
            continue

        if skip_depth > 0:
            # Inside a skipped (inaccessible) struct – track nesting only
            mm = _RE_MEMBER_LINE.match(line)
            if mm:
                rest = mm.group(4)
                type_str = rest.split("//")[0].rstrip().rstrip(";").strip()
                if ":=" in type_str:
                    type_str = type_str.split(":=")[0].strip()
                if type_str.lower() == "struct":
                    skip_depth += 1
            continue

        mm = _RE_MEMBER_LINE.match(line)
        if not mm:
            continue

        member_name = mm.group(2)
        attrs_str = mm.group(3) or ""
        rest = mm.group(4)

        attrs = _parse_db_attrs(attrs_str)

        # Comment
        comment = ""
        if "//" in rest:
            rest_part, comment = rest.split("//", 1)
            comment = comment.strip()
            rest = rest_part

        rest = rest.rstrip().rstrip(";").strip()

        # Start value
        start_value = ""
        if ":=" in rest:
            type_str, sv = rest.split(":=", 1)
            type_str = type_str.strip()
            start_value = sv.strip()
        else:
            type_str = rest.strip()

        is_struct = type_str.lower() == "struct"
        display_dt = "Struct" if is_struct else format_datatype(type_str)

        full_path = [block_name] + [s[0] for s in path_stack] + [member_name]

        if not attrs["ExternalAccessible"]:
            if is_struct:
                skip_depth = 1
            continue

        results.append({
            "Name": ".".join(full_path),
            "Data type": display_dt,
            "Start value": start_value,
            "Retain": base_remanence == "Retain",
            "Accessible from HMI/OPC UA": attrs["ExternalAccessible"],
            "Writable from HMI/OPC": attrs["ExternalWritable"],
            "Visible in HMI engineering": attrs["ExternalVisible"],
            "Setpoint": attrs["SetPoint"],
            "Comment": comment,
        })

        if is_struct:
            path_stack.append((member_name, attrs))
        elif is_udt_reference(type_str):
            udt_name = normalize_udt_name(type_str)
            if udt_name and udt_name in udt_cache:
                _expand_udt_for_db(udt_cache[udt_name], full_path,
                                   base_remanence, udt_cache, results)

    return results


def _expand_udt_for_db(udt_members, parent_path, remanence, udt_cache, results):
    """Expand UDT members recursively for a .db Global DB variable."""
    for m_info in udt_members.values():
        _walk_udt_from_cache(m_info, parent_path, remanence, udt_cache, results, 0)


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel Output
# ═══════════════════════════════════════════════════════════════════════════════

def write_excel(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "HMI Variables"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    bool_cols = {4, 5, 6, 7, 8}  # 1-indexed columns with boolean values

    for row_idx, item in enumerate(results, 2):
        for col_idx, key in enumerate(EXCEL_HEADERS, 1):
            val = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx in bool_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    # Auto-size columns (sample first 500 rows)
    for col_idx, h in enumerate(EXCEL_HEADERS, 1):
        max_len = len(h)
        for row_idx in range(2, min(len(results) + 2, 500)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 3, 80
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_file)
    print(f"  Saved {len(results)} rows → {output_file}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
#  Auto-detection of PLC folder
# ═══════════════════════════════════════════════════════════════════════════════

def find_plc_base(start_dir):
    """Search for a PLC folder containing 'PLC data types' and 'Program blocks'.
    Walks into TiaExport/Projects/*/Devices/PLCs/* under start_dir.
    Returns the first matching absolute path or None."""
    tia_export = os.path.join(start_dir, "TiaExport")
    search_roots = [tia_export, start_dir] if os.path.isdir(tia_export) else [start_dir]

    for search_root in search_roots:
        for root, dirs, _ in os.walk(search_root):
            if "PLC data types" in dirs and "Program blocks" in dirs:
                return root
            # Prune deep trees — PLC folder is typically within 7 levels
            depth = root[len(search_root):].count(os.sep)
            if depth > 7:
                dirs.clear()
    return None


def resolve_paths(plc_base_arg, output_arg):
    """Resolve PLC base dir, UDT dir, Program blocks dir, and output file."""
    if plc_base_arg:
        plc_base = os.path.abspath(plc_base_arg)
    else:
        print("Auto-detecting PLC folder ...")
        plc_base = find_plc_base(os.getcwd())
        if plc_base is None:
            sys.exit(
                "ERROR: Could not find a PLC folder (with 'PLC data types' and "
                "'Program blocks' subdirectories).\n"
                "Run from the project root or pass the PLC folder path as argument:\n"
                "  python _extractHMI.py <path-to-PLC-folder>"
            )
        print(f"  Found: {plc_base}")

    udt_dir = os.path.join(plc_base, "PLC data types")
    prog_dir = os.path.join(plc_base, "Program blocks")

    for d, label in [(plc_base, "PLC base"), (udt_dir, "PLC data types"), (prog_dir, "Program blocks")]:
        if not os.path.isdir(d):
            sys.exit(f"ERROR: {label} directory not found: {d}")

    if output_arg:
        output_file = os.path.abspath(output_arg)
    else:
        plc_name = os.path.basename(plc_base)
        output_file = os.path.join(os.getcwd(), "UserFiles", f"HMI_Variables_{plc_name}.xlsx")

    return plc_base, udt_dir, prog_dir, output_file


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract HMI-accessible variables from TIA Portal DB/IDB exports to Excel.",
    )
    parser.add_argument(
        "plc_base", nargs="?", default=None,
        help="Path to the PLC folder containing 'PLC data types' and 'Program blocks'. "
             "If omitted, auto-detects from TiaExport/ under the current directory.",
    )
    parser.add_argument(
        "-o", "--output", default=None,
        help="Output Excel file path. Default: HMI_Variables_<project>.xlsx in cwd.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    plc_base, udt_dir, prog_dir, output_file = resolve_paths(args.plc_base, args.output)

    print("Building UDT cache ...")
    udt_cache = build_udt_cache(udt_dir)
    print(f"  {len(udt_cache)} UDT definitions loaded")

    print("Building FB interface cache ...")
    fb_cache, xml_fb = build_fb_cache(prog_dir)
    print(f"  {len(fb_cache)} FB/FC/OB interfaces loaded ({xml_fb} XML + {len(fb_cache) - xml_fb} s7dcl)")

    print("Scanning DB/IDB files ...")
    all_results = []
    xml_count = 0
    db_count = 0

    for filepath in find_xml_files(prog_dir):
        rows = process_xml_file(filepath, udt_cache, fb_cache)
        if rows:
            xml_count += 1
            all_results.extend(rows)

    for filepath in find_db_files(prog_dir):
        rows = process_db_file(filepath, udt_cache)
        if rows:
            db_count += 1
            all_results.extend(rows)

    all_results.sort(key=lambda x: x["Name"])

    unique_blocks = {r["Name"].split(".")[0] for r in all_results}
    print(f"  {xml_count} XML + {db_count} .db files with HMI variables")
    print(f"  {len(all_results)} total HMI-accessible rows from {len(unique_blocks)} blocks")

    write_excel(all_results, output_file)


if __name__ == "__main__":
    main()
